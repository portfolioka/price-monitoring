import openpyxl
import datetime

FILEPATH = '' #price_manage.xlsxのフルパスを''内に記載する
SCRAP_LIST_SHEET = '対象サイト'
PRICE_LOG_SHEET = '価格管理表'
ERROR_LOG_SHEET = 'エラーログ'

def get_info(sheet) -> list:
    """[Excelの指定sheetから2行目以降の行情報を取得する]

    Args:
        sheet ([str]): [Excelのシート名]

    Returns:
        list: [2行目以降の行情報リスト]
    """
    wb = openpyxl.load_workbook(FILEPATH)
    return list(wb[sheet].iter_rows(min_row=2))

def write(sheet, content, id, name):
    """[価格やエラーログなどを書き込む]"""
    wb = openpyxl.load_workbook(FILEPATH)
    last_row_num = wb[sheet].max_row +1
    wb[sheet].cell(row=last_row_num, column=1).value = id
    wb[sheet].cell(row=last_row_num, column=2).value = name
    wb[sheet].cell(row=last_row_num, column=3).value = content
    wb[sheet].cell(row=last_row_num, column=4).value = datetime.date.today().strftime('%Y/%-m/%-d')
    wb.save(FILEPATH)

def get_cell_value_list(id, row_list) -> list:
    """[指定した商品idの価格をリストへ格納し、リストを返す]

    Args:
        id ([int]): [商品id]
        row_list ([list]): [価格記録sheetの2行以降のデータ]

    Returns:
        list: [指定した商品idの価格リスト]
    """
    return [cell[2].value for cell in row_list if cell[0].value == id]